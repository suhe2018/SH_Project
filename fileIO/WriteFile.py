# -*- coding: utf-8 -*-
# 读写2003 excel
import xlrd
import xlwt
# 读写2007 excel
import openpyxl


def write03Excel(path):
	wb = xlwt.Workbook()
	sheet = wb.add_sheet("2003测试表")
	value = [["名称", "价格", "出版社", "语言"],
	         ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
	         ["暗时间", "32.4", "人民邮电出版社", "中文"],
	         ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
	for i in range(0, 4):
		for j in range(0, len(value[i])):
			sheet.write(i, j, value[i][j])
	wb.save(path)
	print("写入数据成功！")


def read03Excel(path):
	workbook = xlrd.open_workbook(path)
	sheets = workbook.sheet_names()
	worksheet = workbook.sheet_by_name(sheets[0])
	for i in range(0, worksheet.nrows):
		row = worksheet.row(i)
		for j in range(0, worksheet.ncols):
			print(worksheet.cell_value(i, j), "\t", end="")
		print()


def write07Excel(path):
	wb = openpyxl.Workbook()
	sheet = wb.active
	sheet.title = '2007测试表'
	
	value = [["名称", "价格", "出版社", "语言"],
	         ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
	         ["暗时间", "32.4", "人民邮电出版社", "中文"],
	         ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
	for i in range(0, 4):
		for j in range(0, len(value[i])):
			sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
	
	wb.save(path)
	print("写入数据成功！")

a=[]
b=[]
def read07Excel(path):
	wb = openpyxl.load_workbook(path)
	sheet = wb.get_sheet_by_name('2007测试表')
	
	for row in sheet.rows:
		for cell in row:
			if (cell.column=='A'):
				a.append(cell.value)
			else:
				b.append(cell.value)



file_2003 = 'data/2003.xls'
file_2007 = 'data/2007.xlsx'


# read07Excel(file_2007)

import pandas as pd
#任意的多组列表
# a=[1,2,3]
# b=[3,3,5]
#字典中的key值即为csv中列名
dataframe = pd.DataFrame({'a_name':a,'b_name':b})

#将DataFrame存储为csv,index表示是否显示行名，default=True
# dataframe.to_csv("test.csv",index=False,sep=',')

#读csv
df=pd.read_csv('test.csv',header=None,sep=' ')
print (df.columns[0])
print (df.values[0:1])

#写入txt
from pyhanlp import *
import codecs

# print(HanLP.parseDependency("徐先生还具体帮助他确定了把画雄鹰、松鼠和麻雀作为主攻目标。"))
with codecs.open("douban.txt","w","utf-8-sig") as f:
	str=str(HanLP.parseDependency("徐先生还具体帮助他确定了把画雄鹰、松鼠和麻雀作为主攻目标。"))
	f.write(str)
	
